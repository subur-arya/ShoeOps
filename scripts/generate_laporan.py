#!/usr/bin/env python3
"""Generate laporan ShoeOps XLSX yang rapi dengan openpyxl."""

import sys
import json
from datetime import datetime, date
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Warna ──────────────────────────────────────────────────────
C_BRAND      = "D4510C"   # orange brand
C_NAVY       = "1E3A5F"   # navy biru untuk header utama
C_NAVY_LIGHT = "2E5F9E"   # biru medium untuk table header
C_TITLE_BG   = "EBF3FB"   # biru sangat muda untuk title block
C_SECTION_BG = "FEF9F5"   # krem sangat muda untuk section
C_HEADER_TBL = "D6E4F7"   # biru muda untuk table header
C_ALT_ROW    = "F7FAFD"   # biru pucat alt row
C_GREEN      = "1A7A3D"   # hijau tua
C_GREEN_BG   = "E8F5EE"   # hijau muda bg
C_RED        = "C0392B"   # merah tua
C_RED_BG     = "FDEDEC"   # merah muda bg
C_BLUE       = "1A5EA8"   # biru tua
C_ORANGE     = "C0540A"   # orange tua
C_GRAY       = "6B7280"   # abu netral
C_GRAY_LIGHT = "9CA3AF"   # abu muda
C_WHITE      = "FFFFFF"
C_BORDER     = "BFCFE8"   # border biru muda
C_BORDER_SEC = "E2EAF4"   # border sangat muda

def px(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def ft(bold=False, color=C_NAVY, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bd(all_sides=C_BORDER, thick_bottom=False):
    s  = Side(style="thin",   color=all_sides)
    sb = Side(style="medium", color=C_NAVY_LIGHT) if thick_bottom else s
    return Border(left=s, right=s, top=s, bottom=sb)

def bd_outer():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def bd_none():
    return Border()

def fmt_rp(n):
    try:
        return f"Rp {int(n):,}".replace(",", ".")
    except:
        return "Rp 0"

def fmt_pct(n):
    try:
        return f"{float(n):.1f}%"
    except:
        return "0.0%"

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title_block(ws, title, subtitle, row_start):
    """Tulis blok judul besar di atas sheet."""
    ws.merge_cells(f"A{row_start}:H{row_start}")
    c = ws[f"A{row_start}"]
    c.value     = f"  {title}"
    c.font      = ft(bold=True, color=C_NAVY, size=14)
    c.fill      = px(C_TITLE_BG)
    c.alignment = al("left", "center")
    c.border    = Border(
        left=Side(style="thick", color=C_BRAND),
        top=Side(style="thin", color=C_BORDER),
        right=Side(style="thin", color=C_BORDER),
    )

    ws.merge_cells(f"A{row_start+1}:H{row_start+1}")
    c2 = ws[f"A{row_start+1}"]
    c2.value     = f"  {subtitle}"
    c2.font      = ft(color=C_GRAY, size=9, italic=True)
    c2.fill      = px(C_TITLE_BG)
    c2.alignment = al("left", "center")
    c2.border    = Border(
        left=Side(style="thick", color=C_BRAND),
        bottom=Side(style="thin", color=C_BORDER),
        right=Side(style="thin", color=C_BORDER),
    )

    ws.row_dimensions[row_start].height   = 30
    ws.row_dimensions[row_start+1].height = 18
    return row_start + 3

def write_section_header(ws, label, row, ncols=8):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws[f"A{row}"]
    c.value     = f"  ▸  {label}"
    c.font      = ft(bold=True, color=C_NAVY, size=9)
    c.fill      = px(C_HEADER_TBL)
    c.alignment = al("left", "center")
    c.border    = Border(
        left=Side(style="medium", color=C_NAVY_LIGHT),
        top=Side(style="thin", color=C_BORDER),
        bottom=Side(style="thin", color=C_BORDER),
        right=Side(style="thin", color=C_BORDER),
    )
    ws.row_dimensions[row].height = 20
    return row + 1

def write_table_header(ws, headers, row, fills=None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = ft(bold=True, color=C_WHITE, size=9)
        c.fill      = px(C_NAVY_LIGHT)
        c.alignment = al("center", "center")
        c.border    = bd()
    ws.row_dimensions[row].height = 20
    return row + 1

def write_data_row(ws, values, row, alt=False):
    bg = C_ALT_ROW if alt else C_WHITE
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = ft(size=9, color=C_NAVY)
        c.fill      = px(bg)
        c.alignment = al("left" if col <= 2 else "right" if isinstance(v, (int, float)) else "left")
        c.border    = bd(C_BORDER_SEC)
    ws.row_dimensions[row].height = 16

def write_kv_row(ws, label, value, row, value_color=None):
    cl = ws.cell(row=row, column=1, value=label)
    cl.font = ft(color=C_GRAY_LIGHT, size=9)
    cl.fill = px(C_WHITE)
    vc = ws.cell(row=row, column=2, value=value)
    vc.font      = ft(bold=True, color=value_color or C_NAVY, size=10)
    vc.alignment = al("left")
    vc.fill      = px(C_WHITE)
    ws.row_dimensions[row].height = 18

# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

def main(input_path, output_path):
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    orders   = data.get("orders",   [])
    expenses = data.get("expenses", [])
    now_str  = data.get("now",      datetime.now().isoformat())
    now      = datetime.fromisoformat(now_str[:19])

    # Bulan yang tersedia (maks 24)
    months_set = set()
    for o in orders:
        months_set.add(o["created_at"][:7])
    for e in expenses:
        months_set.add(e["date"][:7])

    # 24 bulan terakhir
    all_months = []
    for i in range(24):
        d = date(now.year, now.month, 1)
        m = (d.month - 1 - i) % 12 + 1
        y = d.year + (d.month - 1 - i) // 12 * (-1 if (d.month - 1 - i) < 0 else 0)
        # Simpler approach
        import calendar
        target = date(now.year, now.month, 1)
        # Go back i months
        month = target.month - i
        year  = target.year
        while month <= 0:
            month += 12
            year  -= 1
        all_months.append(f"{year:04d}-{month:02d}")

    months_with_data = [m for m in all_months if m in months_set]

    wb = Workbook()
    wb.remove(wb.active)  # hapus sheet default

    # ── SHEET ANALITIK KESELURUHAN ──────────────────────────
    ws_all = wb.create_sheet("Analitik Keseluruhan")
    ws_all.sheet_view.showGridLines = False
    ws_all.freeze_panes = "A4"

    row = write_title_block(ws_all,
        "LAPORAN ANALITIK KESELURUHAN — ShoeOps",
        f"Data 24 bulan terakhir  ·  Dibuat: {now.strftime('%d %B %Y %H:%M')}",
        1)

    # Ringkasan total
    total_rev  = sum(o["total_price"] for o in orders)
    total_exp  = sum(e["amount"]      for e in expenses)
    total_laba = total_rev - total_exp
    total_ord  = len(orders)

    row = write_section_header(ws_all, "RINGKASAN KESELURUHAN", row, 6)
    for label, val, color in [
        ("Total Pesanan",     total_ord,                 None     ),
        ("Total Pemasukan",   fmt_rp(total_rev),         C_GREEN  ),
        ("Total Pengeluaran", fmt_rp(total_exp),         C_RED    ),
        ("Laba Bersih",       fmt_rp(total_laba),        C_BRAND  ),
        ("Margin Laba",       fmt_pct(total_laba / total_rev * 100 if total_rev else 0), C_BLUE),
    ]:
        write_kv_row(ws_all, label, val, row, color)
        row += 1
    row += 1

    # Tabel per bulan
    row = write_section_header(ws_all, "KINERJA PER BULAN", row, 6)
    row = write_table_header(ws_all,
        ["Bulan", "Pesanan", "Pemasukan (Rp)", "Pengeluaran (Rp)", "Laba (Rp)", "Margin (%)"],
        row)
    MONTH_ID = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]
    for idx, ym in enumerate(all_months):
        mo = [o for o in orders   if o["created_at"][:7] == ym]
        me = [e for e in expenses if e["date"][:7]       == ym]
        if not mo and not me: continue
        rev  = sum(o["total_price"] for o in mo)
        exp  = sum(e["amount"]      for e in me)
        laba = rev - exp
        mg   = round(laba / rev * 100, 1) if rev else 0
        y, m = ym.split("-")
        label = f"{MONTH_ID[int(m)-1]} {y}"
        write_data_row(ws_all,
            [label, len(mo), rev, exp, laba, f"{mg}%"],
            row, alt=(idx % 2 == 0))
        row += 1
    row += 2

    # Treatment terlaris
    treat_map = defaultdict(lambda: {"count": 0, "rev": 0})
    for o in orders:
        for item in o.get("order_items", []):
            name = item.get("treatment_name", "").split("||")[0]
            treat_map[name]["count"] += item.get("quantity", 1)
            treat_map[name]["rev"]   += item.get("price", 0) * item.get("quantity", 1)
    top_treats = sorted(treat_map.items(), key=lambda x: x[1]["count"], reverse=True)[:20]

    row = write_section_header(ws_all, "TREATMENT TERLARIS (TOP 20)", row, 4)
    row = write_table_header(ws_all, ["Treatment", "Terjual", "Pendapatan (Rp)", "Avg per Order (Rp)"], row)
    for idx, (name, v) in enumerate(top_treats):
        avg = v["rev"] // v["count"] if v["count"] else 0
        write_data_row(ws_all, [name, v["count"], v["rev"], avg], row, alt=(idx % 2 == 0))
        row += 1
    row += 2

    # Customer terbaik
    cust_map = defaultdict(lambda: {"name": "", "orders": 0, "total": 0})
    for o in orders:
        cid  = o.get("customer_id", "")
        name = (o.get("customers") or {}).get("name", "")
        cust_map[cid]["name"]   = name
        cust_map[cid]["orders"] += 1
        cust_map[cid]["total"]  += o["total_price"]
    top_custs = sorted(cust_map.values(), key=lambda x: x["total"], reverse=True)[:20]

    row = write_section_header(ws_all, "CUSTOMER TERBAIK (TOP 20)", row, 4)
    row = write_table_header(ws_all, ["Customer", "Pesanan", "Total Belanja (Rp)", "Rata-rata/Order (Rp)"], row)
    for idx, c in enumerate(top_custs):
        avg = c["total"] // c["orders"] if c["orders"] else 0
        write_data_row(ws_all, [c["name"], c["orders"], c["total"], avg], row, alt=(idx % 2 == 0))
        row += 1

    set_col_widths(ws_all, [18, 12, 18, 18, 16, 12, 12, 12])

    # ── SHEET PER BULAN ─────────────────────────────────────
    for ym in months_with_data:
        y, m = ym.split("-")
        label = f"{MONTH_ID[int(m)-1]} {y}"

        mo = [o for o in orders   if o["created_at"][:7] == ym]
        me = [e for e in expenses if e["date"][:7]       == ym]

        ws = wb.create_sheet(label)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        rev  = sum(o["total_price"] for o in mo)
        exp  = sum(e["amount"]      for e in me)
        laba = rev - exp

        row = write_title_block(ws,
            f"LAPORAN BULANAN — {label.upper()}",
            f"ShoeOps  ·  Dibuat: {now.strftime('%d %B %Y %H:%M')}",
            1)

        # Ringkasan bulan
        row = write_section_header(ws, "RINGKASAN BULAN INI", row, 8)
        items_kv = [
            ("Total Pesanan",     str(len(mo)),          None   ),
            ("Pemasukan",         fmt_rp(rev),           C_GREEN),
            ("Pengeluaran",       fmt_rp(exp),           C_RED  ),
            ("Laba Bersih",       fmt_rp(laba),          C_BRAND),
            ("Margin",            fmt_pct(laba / rev * 100 if rev else 0), C_BLUE),
        ]
        for label_kv, val, color in items_kv:
            write_kv_row(ws, label_kv, val, row, color)
            row += 1
        row += 1

        # Detail pesanan
        row = write_section_header(ws, f"DETAIL PESANAN ({len(mo)} transaksi)", row, 8)
        if mo:
            row = write_table_header(ws,
                ["Kode", "Customer", "No HP", "Treatment(s)", "Total (Rp)", "Status", "Tanggal", "Bayar"],
                row)
            for idx, o in enumerate(sorted(mo, key=lambda x: x["created_at"])):
                treats = ", ".join(
                    i.get("treatment_name","").split("||")[0]
                    for i in o.get("order_items", [])
                ) or "-"
                write_data_row(ws, [
                    o.get("order_code", ""),
                    (o.get("customers") or {}).get("name", ""),
                    (o.get("customers") or {}).get("phone", ""),
                    treats,
                    o.get("total_price", 0),
                    o.get("status", ""),
                    o.get("created_at", "")[:10],
                    o.get("payment_method", "") or "",
                ], row, alt=(idx % 2 == 0))
                row += 1

            # Subtotal
            ws.cell(row=row, column=4, value="TOTAL").font = ft(bold=True, size=9)
            c = ws.cell(row=row, column=5, value=rev)
            c.font      = ft(bold=True, color=C_GREEN, size=9)
            c.alignment = al("right")
            c.fill      = px(C_GREEN_BG)
            ws.row_dimensions[row].height = 16
            row += 2
        else:
            ws.cell(row=row, column=1, value="Tidak ada pesanan bulan ini").font = ft(italic=True, color=C_GRAY, size=9)
            row += 2

        # Detail pengeluaran
        row = write_section_header(ws, f"DETAIL PENGELUARAN ({len(me)} item)", row, 8)
        if me:
            row = write_table_header(ws,
                ["Nama Pengeluaran", "Kategori", "Jumlah (Rp)", "Tanggal", "Catatan", "", "", ""],
                row)
            for idx, e in enumerate(sorted(me, key=lambda x: x["date"])):
                cat = (e.get("category") or {}).get("name", "Tanpa Kategori")
                write_data_row(ws, [
                    e.get("name", ""),
                    cat,
                    e.get("amount", 0),
                    e.get("date", ""),
                    e.get("notes", "") or "",
                ], row, alt=(idx % 2 == 0))
                row += 1

            # Subtotal
            ws.cell(row=row, column=2, value="TOTAL").font = ft(bold=True, size=9)
            c = ws.cell(row=row, column=3, value=exp)
            c.font      = ft(bold=True, color=C_RED, size=9)
            c.alignment = al("right")
            c.fill      = px(C_RED_BG)
            ws.row_dimensions[row].height = 16
            row += 2
        else:
            ws.cell(row=row, column=1, value="Tidak ada pengeluaran bulan ini").font = ft(italic=True, color=C_GRAY, size=9)
            row += 2

        set_col_widths(ws, [14, 20, 14, 32, 14, 12, 12, 14])

    # Pindah Analitik ke posisi pertama
    wb.move_sheet("Analitik Keseluruhan", offset=-len(wb.sheetnames)+1)

    wb.save(output_path)
    print(f"OK: {output_path}")

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])